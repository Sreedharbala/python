#day 16 feb12

from openpyxl import load_workbook

#load the workbook
workbook=load_workbook('C:\\Users\\Lenovo\\Documents\\one.xlsx')

#extract data from sheet 1
sheet1=workbook['Sheet1']
data_sheet1=[]
for row in sheet1.iter_rows():
    data_row=[]
    for cell in row:
        data_row.append(cell.value)
    data_sheet1.append(data_row)


#extraxt data from sheet2
# sheet2=workbook['sheet2']
# data_sheet2=[]
#
# for row in sheet2.iter_rows(values_only=True):
#     data_sheet2.append(row)
#
# #printing the data from sheet1
print("Data from Sheet1:")
for row in data_sheet1:
     print(row)
#
# #printing the data from sheet2
# print("Data from sheet2:")
# for row in data_sheet2:
#     print(row)