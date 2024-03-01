#day15 feb 08
#method overloading:it is has languages like java c++ refers to the ability to define multiple methods with a same name,
#but different parameters with un a class.However ,python does not support method overloading in the same way that java or c++ does.

#in python,you can only have one method with agiven name in class .if you define multiple methods with the same
#name,the last method definition will override the previous ones.Therefore,method overloading,as commonly understood
#does not exist in python.

# class mathoperations:
#     def add(self,a,b):
#         return a+b
#     def add(self,a,b,c):
#         return a+b+c
# math1=mathoperations()
# result1=math1.add(2,3,4)
# print("Result:",result1)#output:9
#
# resutl2=math1.add(2,3,10 )

#to open xl sheet in pyhton
#import openpyxl:
#go to the location of xl path
#book=openpyxl.load_workbook(c:\\users\admin\desktop\xlpy.xlsx")
#to open the xl sheet
#sheet=book.active

#to identify the rows and columns in the xl sheet
#cell=sheet.cell(row=2,column=4)
#to print the values of the cell
#print(cell.value)


import openpyxl
book=openpyxl.load_workbook("C:\\Users\\Lenovo\\Documents\\Book1.xlsx")
sheet=book.active

cell=sheet.cell(row = 2 , column = 2)
print(cell.value)